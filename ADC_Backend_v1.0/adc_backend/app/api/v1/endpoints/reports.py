from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from io import BytesIO

from app.db.session import get_db
from app.models.domain import Domain
from app.models.application import Application
from app.models.score import Score
from app.models.pricing import Pricing
from app.models.query import Query
from app.models.organisation import Organisation
from app.core.dependencies import get_current_org_id

router = APIRouter()

async def _get_portfolio_data(org_id: int, db: AsyncSession) -> dict:
    org_q = await db.execute(select(Organisation).where(Organisation.id == org_id))
    org = org_q.scalar_one()
    domains_q = await db.execute(select(Domain).where(Domain.org_id == org_id, Domain.is_active == True).order_by(Domain.sort_order))
    domains = domains_q.scalars().all()
    apps_q = await db.execute(
        select(Application).join(Domain).where(Domain.org_id == org_id, Application.is_active == True, Domain.is_active == True)
    )
    applications = apps_q.scalars().all()
    scores_q = await db.execute(
        select(Score).join(Application).join(Domain).where(Domain.org_id == org_id)
    )
    all_scores = scores_q.scalars().all()
    scores_map = {}
    for s in all_scores:
        scores_map.setdefault(s.app_id, {})[s.criterion_index] = s.score
    pricing_q = await db.execute(
        select(Pricing).join(Application).join(Domain).where(Domain.org_id == org_id)
    )
    pricing_map = {p.app_id: p for p in pricing_q.scalars().all()}
    queries_q = await db.execute(select(Query).where(Query.org_id == org_id, Query.status != "Resolved"))
    open_queries = queries_q.scalars().all()
    return dict(org=org, domains=domains, applications=applications,
                scores_map=scores_map, pricing_map=pricing_map, open_queries=open_queries)

@router.get("/summary")
async def portfolio_summary(org_id: int = Depends(get_current_org_id), db: AsyncSession = Depends(get_db)):
    data = await _get_portfolio_data(org_id, db)
    domain_map = {d.id: d for d in data["domains"]}
    result = {"organisation": data["org"].name, "domains": [], "applications": [], "open_queries": []}
    for domain in data["domains"]:
        domain_apps = [a for a in data["applications"] if a.domain_id == domain.id]
        scores_list = []
        for a in domain_apps:
            app_scores = data["scores_map"].get(a.id, {})
            if app_scores:
                scores_list.append(sum(app_scores.values()) / len(app_scores))
        avg = round(sum(scores_list) / len(scores_list), 2) if scores_list else None
        result["domains"].append({"id": domain.id, "name": domain.name, "app_count": len(domain_apps), "avg_score": avg})
    for app in data["applications"]:
        app_scores = data["scores_map"].get(app.id, {})
        avg = round(sum(app_scores.values()) / len(app_scores), 2) if app_scores else None
        pricing = data["pricing_map"].get(app.id)
        tco = float(pricing.licence_cost + pricing.maintenance_cost + pricing.implementation_cost) if pricing else 0
        result["applications"].append({
            "id": app.id, "name": app.name, "vendor": app.vendor,
            "domain": domain_map.get(app.domain_id, {}).name if app.domain_id in domain_map else None,
            "scores": app_scores, "avg_score": avg,
            "tco": tco, "recommendation": pricing.recommendation if pricing else None,
        })
    for q in data["open_queries"]:
        result["open_queries"].append({"id": q.id, "title": q.title, "status": q.status, "priority": q.priority, "assignee": q.assignee, "due_date": str(q.due_date) if q.due_date else None})
    return result

@router.get("/export/excel")
async def export_excel(org_id: int = Depends(get_current_org_id), db: AsyncSession = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    data = await _get_portfolio_data(org_id, db)
    wb = openpyxl.Workbook()
    # ── Sheet 1: Applications ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Applications"
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    headers = ["Application", "Vendor", "Domain", "Business Fit", "Technical Health", "Cost Efficiency", "Risk Level", "Avg Score", "Licence (R)", "Maint. (R)", "Impl. (R)", "Total TCO (R)", "Recommendation"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    domain_map = {d.id: d.name for d in data["domains"]}
    for app in data["applications"]:
        app_scores = data["scores_map"].get(app.id, {})
        avg = round(sum(app_scores.values()) / len(app_scores), 2) if app_scores else ""
        p = data["pricing_map"].get(app.id)
        ws.append([
            app.name, app.vendor or "", domain_map.get(app.domain_id, ""),
            app_scores.get(0, ""), app_scores.get(1, ""), app_scores.get(2, ""), app_scores.get(3, ""), avg,
            float(p.licence_cost) if p else 0, float(p.maintenance_cost) if p else 0,
            float(p.implementation_cost) if p else 0,
            float(p.licence_cost + p.maintenance_cost + p.implementation_cost) if p else 0,
            str(p.recommendation or "") if p else "",
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16
    # ── Sheet 2: Queries ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Open Queries")
    ws2.append(["Title", "Status", "Priority", "Category", "Assignee", "Due Date"])
    for cell in ws2[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
    for q in data["open_queries"]:
        ws2.append([q.title, q.status, q.priority, q.category, q.assignee or "", str(q.due_date or "")])
    # ── Sheet 3: Info ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Info")
    ws3.append(["Organisation", data["org"].name])
    ws3.append(["Department", data["org"].department or ""])
    ws3.append(["Reference", data["org"].reference_no or ""])
    ws3.append(["Financial Year", data["org"].financial_year or ""])
    ws3.append(["Export Date", str(__import__("datetime").date.today())])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"ADC_Report_{data['org'].name.replace(' ', '_')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})
